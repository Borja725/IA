import pandas as pd
from tabulate import tabulate
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import re

# Define file paths
INPUT_FILE = 'verbatim_predictions_llama3-70b_v1_prompt_engin.csv'
OUTPUT_FILE = 'verbatim_predictions_llama3-70b_v1_prompt_engin.csv_SERVER.csv'
SUMMARY_IMAGE_FILE = 'RESULTADOS\\verbatim_predictions_llama3-70b_summary_statistics_SERVER.png'
SUMMARY_EXCEL_FILE = 'RESULTADOS\\verbatim_predictions_llama3-70b_summary_statistics_SERVER.xlsx'

# Define boolean variables for output options
PRINT_CONSOLE = True
SAVE_IMAGE = True
SAVE_EXCEL = False

def format_codes(codes):
    """Format a comma-separated string of codes to be four digits each, handling brackets."""
    if pd.isna(codes):
        return ''
    
    # Remove brackets and any surrounding whitespace
    codes = re.sub(r'[^\w,]', '', str(codes))
    
    formatted_codes = []
    for code in str(codes).split(','):
        code = code.strip()
        if code.isdigit():  # Check if the code is numeric
            formatted_codes.append(f'{int(code):04}')
        else:
            formatted_codes.append(code)  # Keep non-numeric codes as they are
    return ','.join(formatted_codes)

def is_text_present(codes):
    """Check if the string contains non-numeric text."""
    return any(not code.isdigit() for code in str(codes).split(','))

def normalize_codes(codes):
    """Normalize codes by removing extra commas, spaces, and brackets."""
    if pd.isna(codes):
        return ''
    
    # Remove brackets and any surrounding whitespace
    codes = re.sub(r'[^\w,]', '', str(codes))
    
    normalized = []
    for code in str(codes).split(','):
        code = code.strip()
        if code:  # Ignore empty strings
            if code.isdigit():
                normalized.append(f'{int(code):04}')
            else:
                normalized.append(code)
    return ','.join(normalized)

def compare_codes(code, predictions):
    """Compare code with predictions and return comparison value."""
    normalized_code = normalize_codes(code)
    normalized_predictions = normalize_codes(predictions)

    # If the PREDICTIONS column is empty or contains '[]'
    if pd.isna(predictions) or predictions.strip() == '' or predictions.strip() == '[]':
        return 5

    # If there is text present in either CODES or PREDICTIONS
    if is_text_present(normalized_code) or is_text_present(normalized_predictions):
        return 0

    code_set = set(normalized_code.split(','))
    predictions_set = set(normalized_predictions.split(','))

    # Case 1: Exact match
    if code_set == predictions_set:
        return 1

    # Case 2: Partial positive
    elif code_set.issubset(predictions_set):
        return 2

    # Case 3: Partial negative
    elif (code_set - predictions_set) and (predictions_set <= code_set):
        return 3

    # Case 0: No significant overlap (False Positive)
    else:
        return 0

def process_csv(input_file, output_file):
    """Process the CSV file to compare codes and predictions."""
    # Read the CSV file
    df = pd.read_csv(input_file)

    # Format the CODE and PREDICTIONS columns
    df['CODES'] = df['CODES'].apply(format_codes)
    df['PREDICTIONS'] = df['PREDICTIONS'].apply(format_codes)

    # Add the COMPARISON column
    df['COMPARISON'] = df.apply(lambda row: compare_codes(row['CODES'], row['PREDICTIONS']), axis=1)

    # Save the result to a new CSV file
    df.to_csv(output_file, index=False)

    return df

def print_summary_statistics(df):
    """Print summary statistics."""
    total = len(df)
    total_clasificado = len(df[df['COMPARISON'] != 5])
    no_clasificados = len(df[df['COMPARISON'] == 5])
    tp = len(df[df['COMPARISON'] == 1])
    pp = len(df[df['COMPARISON'] == 2])
    pn = len(df[df['COMPARISON'] == 3])
    fp = len(df[df['COMPARISON'] == 0])  # False positive

    # No clasificados + FP
    no_clasificados_plus_fp = no_clasificados + fp

    # TP + PP + PN
    tp_pp_pn = tp + pp + pn

    # Percentages
    perc_total = {
        "Total": total,
        "Total Clasificado": total_clasificado,
        "No clasificados": no_clasificados,
        "No clas. + FP": no_clasificados_plus_fp,
        "TP+PP+PN": tp_pp_pn,
        "TP (true positive)": tp,
        "PP (Partial positive)": pp,
        "PN (Partial negative)": pn,
        "FP (false positive)": fp
    }

    perc_total_clasificado = {
        "TP (true positive)": (tp / total_clasificado) * 100 if total_clasificado else 0,
        "PP (Partial positive)": (pp / total_clasificado) * 100 if total_clasificado else 0,
        "PN (Partial negative)": (pn / total_clasificado) * 100 if total_clasificado else 0,
        "FP (false positive)": (fp / total_clasificado) * 100 if total_clasificado else 0,
        "No clasificados": (no_clasificados / total_clasificado) * 100 if total_clasificado else 0,
        "No clas. + FP": (no_clasificados_plus_fp / total_clasificado) * 100 if total_clasificado else 0,
        "TP+PP+PN": (tp_pp_pn / total_clasificado) * 100 if total_clasificado else 0
    }

    # Calculate percentages for total
    perc_total_percentage = {
        "Total": 100.0,
        "Total Clasificado": (total_clasificado / total) * 100 if total else 0,
        "No clasificados": (no_clasificados / total) * 100 if total else 0,
        "No clas. + FP": (no_clasificados_plus_fp / total) * 100 if total else 0,
        "TP+PP+PN": (tp_pp_pn / total) * 100 if total else 0,
        "TP (true positive)": (tp / total) * 100 if total else 0,
        "PP (Partial positive)": (pp / total) * 100 if total else 0,
        "PN (Partial negative)": (pn / total) * 100 if total else 0,
        "FP (false positive)": (fp / total) * 100 if total else 0
    }

    # Prepare the data for the table
    table_data = [
        ["Category", "Count", "% sobre el Total", "% sobre total clasificado"],
        ["Total", perc_total["Total"], "", ""],
        ["Total Clasificado", perc_total["Total Clasificado"], f"{perc_total_percentage['Total Clasificado']:.2f}%", ""],
        ["No clasificados", perc_total["No clasificados"], f"{perc_total_percentage['No clasificados']:.2f}%", ""],
        ["No clas. + FP", perc_total["No clas. + FP"], f"{perc_total_percentage['No clas. + FP']:.2f}%", ""],
        ["TP+PP+PN", perc_total["TP+PP+PN"], f"{perc_total_percentage['TP+PP+PN']:.2f}%", f"{perc_total_clasificado['TP+PP+PN']:.2f}%"],
        ["", "", "", ""],
        ["TP (true positive)", perc_total["TP (true positive)"], f"{perc_total_percentage['TP (true positive)']:.2f}%", f"{perc_total_clasificado['TP (true positive)']:.2f}%"],
        ["PP (Partial positive)", perc_total["PP (Partial positive)"], f"{perc_total_percentage['PP (Partial positive)']:.2f}%", f"{perc_total_clasificado['PP (Partial positive)']:.2f}%"],
        ["PN (Partial negative)", perc_total["PN (Partial negative)"], f"{perc_total_percentage['PN (Partial negative)']:.2f}%", f"{perc_total_clasificado['PN (Partial negative)']:.2f}%"],
        ["FP (false positive)", perc_total["FP (false positive)"], f"{perc_total_percentage['FP (false positive)']:.2f}%", f"{perc_total_clasificado['FP (false positive)']:.2f}%"]
    ]

    # Print the table to console if PRINT_CONSOLE is True
    if PRINT_CONSOLE:
        print(tabulate(table_data[1:], headers=table_data[0], tablefmt="grid"))

    # Save the table as an image if SAVE_IMAGE is True
    if SAVE_IMAGE:
        save_table_as_image(table_data, SUMMARY_IMAGE_FILE)

    # Save the table as an Excel file if SAVE_EXCEL is True
    if SAVE_EXCEL:
        save_table_as_excel(table_data, SUMMARY_EXCEL_FILE)



def save_table_as_image(table_data, file_name):
    """Save the table as an image file."""
    fig, ax = plt.subplots(figsize=(12, 4))  # Adjust the figure size as needed
    ax.axis('tight')
    ax.axis('off')

    # Create a color map for alternating row colors
    colors = ['#DCE6F1', '#FFFFFF']  # Light blue and white

    # Create table
    table = ax.table(cellText=table_data[1:], colLabels=table_data[0], cellLoc='center', loc='center')

    # Apply formatting
    for i, key in enumerate(table.get_celld().keys()):
        cell = table.get_celld()[key]
        if key[0] == 0:  # Header row
            cell.set_fontsize(12)
            cell.set_text_props(weight='bold')
        else:  # Data rows
            row_index = key[0]
            cell.set_facecolor(colors[row_index % 2])
            if key[1] == 0:  # First column
                cell.set_text_props(weight='bold')
    
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.2, 1.2)  # Adjust table scale
    plt.savefig(file_name, bbox_inches='tight')
    plt.close()

def save_table_as_excel(table_data, file_name):
    """Save the table as an Excel file with formatted cells."""
    # Convert table data to a DataFrame
    df = pd.DataFrame(table_data[1:], columns=table_data[0])
    
    # Save DataFrame to an Excel file
    df.to_excel(file_name, index=False)

    # Load the workbook and access the sheet
    wb = load_workbook(file_name)
    ws = wb.active

    # Define styles
    bold_font = Font(bold=True)
    fill_color1 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_color2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Apply styles to header row and alternate row colors
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        if row_index == 1:  # Header row
            for cell in row:
                cell.font = bold_font
                cell.fill = fill_color1
        else:  # Data rows
            for cell in row:
                cell.fill = fill_color1 if row_index % 2 == 0 else fill_color2
            row[0].font = bold_font  # Bold for first column

    # Save the updated workbook
    wb.save(file_name)

# Main execution
df = process_csv(INPUT_FILE, OUTPUT_FILE)
print_summary_statistics(df)
